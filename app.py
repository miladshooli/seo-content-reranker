import os
import re
import io
import time
import uuid
import sqlite3
import threading
from datetime import datetime

import requests
from flask import Flask, request, jsonify, render_template, send_file

SERPER_SEARCH_URL = "https://google.serper.dev/search"
SERPER_SCRAPE_URL = "https://scrape.serper.dev"
VOYAGE_RERANK_URL = "https://api.voyageai.com/v1/rerank"
RERANK_MODEL = "rerank-2.5"

DB_PATH = os.environ.get("RERANKER_DB", "/opt/seo-reranker/data.db")
RERANK_BATCH = 100          # docs per Voyage call (token-safe)
MAX_CHUNKS_PER_URL = 200    # cap to control cost; logged when hit

app = Flask(__name__)
from serpiwi_auth import init_auth
init_auth(app, "Serpiwi · ری‌رنکر محتوا")

# ----------------------------------------------------------------------------
# in-memory job registry (results are also persisted to SQLite)
# ----------------------------------------------------------------------------
JOBS = {}
JOBS_LOCK = threading.Lock()


def new_job():
    jid = uuid.uuid4().hex[:12]
    with JOBS_LOCK:
        JOBS[jid] = {"status": "running", "progress": 0, "log": [], "results": None,
                     "error": None, "query_id": None, "keyword": None}
    return jid


def job_update(jid, **kw):
    with JOBS_LOCK:
        j = JOBS.get(jid)
        if not j:
            return
        log_line = kw.pop("log", None)
        if log_line is not None:
            j["log"].append(log_line)
        j.update(kw)


def job_get(jid):
    with JOBS_LOCK:
        j = JOBS.get(jid)
        return dict(j) if j else None


# ----------------------------------------------------------------------------
# external services
# ----------------------------------------------------------------------------
def serper_search(keyword, api_key, gl="ir", hl="fa", num_results=10):
    resp = requests.post(
        SERPER_SEARCH_URL,
        headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
        json={"q": keyword, "gl": gl, "hl": hl, "autocorrect": False},
        timeout=45,
    )
    if not resp.ok:
        raise RuntimeError(f"Serper search {resp.status_code}: {resp.text[:200]}")
    return (resp.json().get("organic") or [])[:num_results]


def serper_scrape(url, api_key):
    resp = requests.post(
        SERPER_SCRAPE_URL,
        headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
        json={"url": url, "includeMarkdown": True},
        timeout=90,
    )
    if not resp.ok:
        raise RuntimeError(f"Serper scrape {resp.status_code}: {resp.text[:200]}")
    data = resp.json()
    return data.get("markdown") or data.get("text") or ""


def extract_chunks(markdown_text):
    """Split markdown into paragraph chunks (headings kept as their own chunk)."""
    chunks = []
    if not markdown_text or not markdown_text.strip():
        return chunks
    current = []
    for raw in markdown_text.split("\n"):
        line = raw.strip()
        if not line or line.startswith("#") or line.startswith("|") or line.startswith("!["):
            if current:
                text = " ".join(current).strip()
                if len(text) > 20:
                    chunks.append(text)
                current = []
            if line.startswith("#"):
                heading = re.sub(r"^#+\s*", "", line).strip()
                if heading:
                    chunks.append(heading)
        else:
            current.append(line)
    if current:
        text = " ".join(current).strip()
        if len(text) > 20:
            chunks.append(text)
    # de-dupe while preserving order
    seen, out = set(), []
    for c in chunks:
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out


def voyage_rerank(query, documents, api_key):
    """Return list of {chunk, score} for all documents (batched, with 429 backoff)."""
    out = []
    for start in range(0, len(documents), RERANK_BATCH):
        batch = documents[start:start + RERANK_BATCH]
        backoff = [5, 10, 20, 30, 45]
        attempt = 0
        while True:
            resp = requests.post(
                VOYAGE_RERANK_URL,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={"query": query, "documents": batch, "model": RERANK_MODEL, "top_k": len(batch)},
                timeout=90,
            )
            if resp.status_code == 429 and attempt < len(backoff):
                time.sleep(backoff[attempt])
                attempt += 1
                continue
            if not resp.ok:
                raise RuntimeError(f"Voyage rerank {resp.status_code}: {resp.text[:200]}")
            break
        for item in resp.json().get("data", []):
            idx = item.get("index")
            if idx is None or idx >= len(batch):
                continue
            out.append({"chunk": batch[idx], "score": float(item.get("relevance_score") or 0.0)})
    out.sort(key=lambda x: x["score"], reverse=True)
    return out


# ----------------------------------------------------------------------------
# database
# ----------------------------------------------------------------------------
def db_conn():
    d = os.path.dirname(DB_PATH)
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    conn = db_conn()
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS queries (
        id INTEGER PRIMARY KEY AUTOINCREMENT, query TEXT NOT NULL, timestamp TEXT NOT NULL)""")
    c.execute("""CREATE TABLE IF NOT EXISTS urls (
        id INTEGER PRIMARY KEY AUTOINCREMENT, query_id INTEGER, rank TEXT, url TEXT,
        title TEXT, snippet TEXT, FOREIGN KEY (query_id) REFERENCES queries (id))""")
    c.execute("""CREATE TABLE IF NOT EXISTS chunks (
        id INTEGER PRIMARY KEY AUTOINCREMENT, url_id INTEGER, chunk_index INTEGER,
        chunk_text TEXT, relevance_score REAL, FOREIGN KEY (url_id) REFERENCES urls (id))""")
    conn.commit()
    conn.close()


def save_analysis(keyword, results_data):
    conn = db_conn()
    c = conn.cursor()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("INSERT INTO queries (query, timestamp) VALUES (?, ?)", (keyword, ts))
    qid = c.lastrowid
    for r in results_data:
        c.execute("INSERT INTO urls (query_id, rank, url, title, snippet) VALUES (?,?,?,?,?)",
                  (qid, str(r["rank"]), r["url"], r["title"], r.get("snippet", "")))
        uid = c.lastrowid
        for i, ch in enumerate(r["chunks"]):
            c.execute("INSERT INTO chunks (url_id, chunk_index, chunk_text, relevance_score) VALUES (?,?,?,?)",
                      (uid, i, ch["chunk"], ch["score"]))
    conn.commit()
    conn.close()
    return qid


def list_history():
    conn = db_conn()
    rows = conn.execute("""
        SELECT q.id, q.query, q.timestamp,
               COUNT(DISTINCT u.id) AS urls, COUNT(c.id) AS chunks
        FROM queries q
        LEFT JOIN urls u ON u.query_id = q.id
        LEFT JOIN chunks c ON c.url_id = u.id
        GROUP BY q.id ORDER BY q.id DESC LIMIT 100""").fetchall()
    conn.close()
    return [{"id": r[0], "query": r[1], "timestamp": r[2], "urls": r[3], "chunks": r[4]} for r in rows]


def load_query_urls(query_id):
    """Return urls (with their chunk texts) for an existing query, grouped by url."""
    conn = db_conn()
    rows = conn.execute("""
        SELECT u.rank, u.url, u.title, u.snippet, c.chunk_text
        FROM urls u JOIN chunks c ON c.url_id = u.id
        WHERE u.query_id = ? ORDER BY u.id, c.chunk_index""", (query_id,)).fetchall()
    conn.close()
    grouped = {}
    for rank, url, title, snippet, chunk_text in rows:
        g = grouped.setdefault(url, {"rank": rank, "title": title, "snippet": snippet, "chunks": []})
        g["chunks"].append(chunk_text)
    return grouped


def summarize(results_data):
    """Attach per-url avg/max/count summary; keep chunks sorted by score."""
    for r in results_data:
        scores = [c["score"] for c in r["chunks"]] or [0.0]
        r["avg_score"] = sum(scores) / len(scores)
        r["max_score"] = max(scores)
        r["chunk_count"] = len(r["chunks"])
    return results_data


# ----------------------------------------------------------------------------
# background jobs
# ----------------------------------------------------------------------------
def run_analyze(jid, keyword, serper_key, voyage_key, gl, hl, top_n, manual_urls):
    try:
        job_update(jid, log=f"جستجوی گوگل برای «{keyword}»…")
        organic = serper_search(keyword, serper_key, gl, hl, num_results=max(top_n, 10))
        targets = []
        for i, r in enumerate(organic[:top_n], 1):
            targets.append({"rank": i, "url": r.get("link", ""),
                            "title": r.get("title", ""), "snippet": r.get("snippet", "")})
        for m in (manual_urls or []):
            if m.get("url"):
                targets.append({"rank": "دستی", "url": m["url"],
                                "title": m.get("title") or "صفحه دستی", "snippet": "URL دستی"})
        if not targets:
            raise RuntimeError("هیچ نتیجه‌ای برای این کیورد پیدا نشد.")
        job_update(jid, log=f"{len(targets)} صفحه برای پردازش انتخاب شد.")

        results = []
        total = len(targets)
        for i, t in enumerate(targets, 1):
            job_update(jid, progress=int((i - 1) / total * 100),
                       log=f"[{i}/{total}] اسکرپ: {t['title'][:55]}")
            try:
                md = serper_scrape(t["url"], serper_key)
            except RuntimeError as e:
                job_update(jid, log=f"   ✗ خطای اسکرپ: {e}")
                continue
            chunks = extract_chunks(md)
            if len(chunks) > MAX_CHUNKS_PER_URL:
                job_update(jid, log=f"   ⚠ {len(chunks)} چانک؛ به {MAX_CHUNKS_PER_URL} محدود شد.")
                chunks = chunks[:MAX_CHUNKS_PER_URL]
            if not chunks:
                job_update(jid, log="   ✗ محتوایی برای چانک‌سازی نبود.")
                continue
            job_update(jid, log=f"   ⟳ رِرَنک {len(chunks)} چانک با Voyage…")
            scored = voyage_rerank(keyword, chunks, voyage_key)
            t["chunks"] = scored
            results.append(t)

        if not results:
            raise RuntimeError("هیچ صفحه‌ای با موفقیت پردازش نشد (اسکرپ/چانک ناموفق).")
        summarize(results)
        qid = save_analysis(keyword, results)
        job_update(jid, status="done", progress=100, results=results, query_id=qid,
                   keyword=keyword, log=f"✓ ذخیره شد (شناسه دیتابیس: {qid}).")
    except Exception as e:
        job_update(jid, status="error", error=str(e), log=f"✗ {e}")


def run_rerank(jid, query_id, new_keyword, voyage_key):
    try:
        job_update(jid, log=f"بارگذاری چانک‌های ذخیره‌شده (کوئری #{query_id})…")
        grouped = load_query_urls(query_id)
        if not grouped:
            raise RuntimeError("دیتایی برای این کوئری در دیتابیس نیست.")
        results = []
        total = len(grouped)
        for i, (url, g) in enumerate(grouped.items(), 1):
            job_update(jid, progress=int((i - 1) / total * 100),
                       log=f"[{i}/{total}] رِرَنک: {g['title'][:55]}")
            scored = voyage_rerank(new_keyword, g["chunks"], voyage_key)
            results.append({"rank": g["rank"], "url": url, "title": g["title"],
                            "snippet": g["snippet"], "chunks": scored})
        summarize(results)
        qid = save_analysis(new_keyword, results)
        job_update(jid, status="done", progress=100, results=results, query_id=qid,
                   keyword=new_keyword, log=f"✓ رِرَنک ذخیره شد (شناسه: {qid}).")
    except Exception as e:
        job_update(jid, status="error", error=str(e), log=f"✗ {e}")


# ----------------------------------------------------------------------------
# routes
# ----------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html",
                           has_serper=bool(os.environ.get("SERPER_KEY")),
                           has_voyage=bool(os.environ.get("VOYAGE_API_KEY")))


@app.route("/api/analyze", methods=["POST"])
def api_analyze():
    p = request.get_json(silent=True) or {}
    keyword = (p.get("keyword") or "").strip()
    serper_key = (p.get("serper_key") or "").strip() or os.environ.get("SERPER_KEY", "")
    voyage_key = (p.get("voyage_key") or "").strip() or os.environ.get("VOYAGE_API_KEY", "")
    gl = (p.get("gl") or "ir").strip()
    hl = (p.get("hl") or "fa").strip()
    try:
        top_n = max(1, min(int(p.get("top_n") or 5), 10))
    except (TypeError, ValueError):
        top_n = 5
    manual_urls = p.get("manual_urls") or []

    if not serper_key:
        return jsonify({"error": "کلید Serper لازم است."}), 400
    if not voyage_key:
        return jsonify({"error": "کلید Voyage لازم است."}), 400
    if not keyword:
        return jsonify({"error": "کیورد نمی‌تواند خالی باشد."}), 400

    jid = new_job()
    job_update(jid, keyword=keyword)
    threading.Thread(target=run_analyze,
                     args=(jid, keyword, serper_key, voyage_key, gl, hl, top_n, manual_urls),
                     daemon=True).start()
    return jsonify({"job_id": jid})


@app.route("/api/rerank", methods=["POST"])
def api_rerank():
    p = request.get_json(silent=True) or {}
    new_keyword = (p.get("keyword") or "").strip()
    voyage_key = (p.get("voyage_key") or "").strip() or os.environ.get("VOYAGE_API_KEY", "")
    query_id = p.get("query_id")

    if not voyage_key:
        return jsonify({"error": "کلید Voyage لازم است."}), 400
    if not new_keyword:
        return jsonify({"error": "کیورد جدید نمی‌تواند خالی باشد."}), 400
    if not query_id:
        return jsonify({"error": "یک کوئری از تاریخچه انتخاب کنید."}), 400

    jid = new_job()
    job_update(jid, keyword=new_keyword)
    threading.Thread(target=run_rerank, args=(jid, query_id, new_keyword, voyage_key),
                     daemon=True).start()
    return jsonify({"job_id": jid})


@app.route("/api/job/<jid>")
def api_job(jid):
    j = job_get(jid)
    if not j:
        return jsonify({"error": "job not found"}), 404
    return jsonify(j)


@app.route("/api/history")
def api_history():
    return jsonify({"queries": list_history()})


@app.route("/api/export/<int:query_id>")
def api_export(query_id):
    from openpyxl import Workbook
    grouped = load_query_urls(query_id)
    conn = db_conn()
    qrow = conn.execute("SELECT query, timestamp FROM queries WHERE id=?", (query_id,)).fetchone()
    # pull scores too (load_query_urls drops them) -> read directly
    rows = conn.execute("""
        SELECT u.rank, u.url, u.title, c.chunk_text, c.relevance_score
        FROM urls u JOIN chunks c ON c.url_id = u.id
        WHERE u.query_id = ? ORDER BY u.id, c.relevance_score DESC""", (query_id,)).fetchall()
    conn.close()
    if not qrow:
        return jsonify({"error": "query not found"}), 404
    keyword = qrow[0]

    wb = Workbook()
    ws = wb.active
    ws.title = "chunks"
    ws.append(["keyword", "rank", "URL", "title", "chunk", "relevance", "chars", "words"])
    for rank, url, title, chunk, score in rows:
        ws.append([keyword, rank, url, title, chunk[:300], round(score, 6), len(chunk), len(chunk.split())])

    ws2 = wb.create_sheet("url_summary")
    ws2.append(["rank", "URL", "title", "avg_relevance", "max_relevance", "chunk_count"])
    agg = {}
    for rank, url, title, chunk, score in rows:
        a = agg.setdefault(url, {"rank": rank, "title": title, "scores": []})
        a["scores"].append(score)
    summ = [(a["rank"], url, a["title"], sum(a["scores"]) / len(a["scores"]),
             max(a["scores"]), len(a["scores"])) for url, a in agg.items()]
    summ.sort(key=lambda x: x[3], reverse=True)
    for rank, url, title, avg, mx, cnt in summ:
        ws2.append([rank, url, title, round(avg, 6), round(mx, 6), cnt])

    ws3 = wb.create_sheet("top_chunks")
    ws3.append(["relevance", "URL", "chunk"])
    for rank, url, title, chunk, score in sorted(rows, key=lambda x: x[4], reverse=True)[:30]:
        ws3.append([round(score, 6), url, chunk[:300]])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    safe = re.sub(r"[^\w؀-ۿ]+", "_", keyword)[:40]
    return send_file(bio, as_attachment=True,
                     download_name=f"reranker_{safe}_{query_id}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/healthz")
def healthz():
    return jsonify({"status": "ok"})


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8001, debug=True)
